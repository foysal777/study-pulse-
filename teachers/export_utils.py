from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


def export_teacher_profiles_to_excel(queryset):
    """
    Export teacher profiles to an Excel file.
    
    Args:
        queryset: QuerySet of TeacherProfile objects
    
    Returns:
        HttpResponse with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Profiles"
    
    # Define header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define data cell style
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Headers
    headers = [
        "ID",
        "Name",
        "Phone Number",
        "Email",
        "Age",
        "Gender",
        "Qualification",
        "Experience",
        "Teaching Medium",
        "Courses/Classes Taught",
        "Other Courses/Classes",
        "Offline Location",
        "Teacher's Room (Link)",
        "Created At",
        "Updated At"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row_idx, teacher_profile in enumerate(queryset, start=2):
        row_data = [
            teacher_profile.id,
            teacher_profile.name,
            teacher_profile.phone_number,
            teacher_profile.user.email,
            teacher_profile.age,
            teacher_profile.gender or "",
            teacher_profile.qualification or "",
            teacher_profile.experience or "",
            teacher_profile.teaching_medium or "",
            teacher_profile.courses_classes_taught or "",
            teacher_profile.other_courses_classes or "",
            teacher_profile.offline_location or "",
            teacher_profile.teachers_room or "",
            teacher_profile.created_at.strftime("%Y-%m-%d %H:%M:%S") if teacher_profile.created_at else "",
            teacher_profile.updated_at.strftime("%Y-%m-%d %H:%M:%S") if teacher_profile.updated_at else "",
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = data_alignment
            cell.border = border
    
    # Adjust column widths
    column_widths = [8, 20, 15, 25, 8, 12, 20, 15, 20, 25, 25, 25, 25, 20, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"teacher_profiles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
